"""
Coffee Shop POS — Database Operations
-------------------------------------
Clean rewrite for the Flask web version.

Responsibilities:
• Handle all database reads/writes
• Track inventory (components)
• Record orders and automatically decrement inventory
• Provide summaries for reports
"""

import sqlite3
import sys
from pathlib import Path
from datetime import datetime
import os
import io, csv, zipfile, time, math
import shutil
from openpyxl import Workbook

# ---------------------------------------------------------------------
# SAFE LOGGING (NO EMOJI / NO print())
# ---------------------------------------------------------------------

def _log_path():
    """Return path to LocalAppData\BrewInsPOS\pos.log"""
    local = os.environ.get("LOCALAPPDATA", "")
    log_dir = Path(local) / "BrewInsPOS"
    log_dir.mkdir(parents=True, exist_ok=True)
    return log_dir / "pos.log"

LOG_FILE = _log_path()

def log(msg: str):
    """Append UTF-8 log messages safely (never crashes)."""
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(f"[{ts}] {msg}\n")
    except:
        pass


# ---------------------------------------------------------------------
# Connection (PyInstaller-safe paths)
# ---------------------------------------------------------------------

# Determine base directory depending on whether we're frozen
if getattr(sys, 'frozen', False):
    # Running inside EXE — store DB in writable LocalAppData
    BASE_DIR = Path(os.environ["LOCALAPPDATA"]) / "BrewInsPOS"
else:
    # Source mode — project root, because this file is backend/db_ops.py
    BASE_DIR = Path(__file__).resolve().parent.parent

DB_FILE = BASE_DIR / "backend" / "pos.db"

# Alias for backup/restore helpers that expect DB_PATH
DB_PATH = str(DB_FILE)


def connect():
    DB_FILE.parent.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON;")
    return conn


def _debug_db_paths():
    log(f"---- DB PATH DEBUG ----")
    log(f"sys.frozen = {getattr(sys, 'frozen', False)}")
    log(f"sys._MEIPASS = {getattr(sys, '_MEIPASS', None)}")
    log(f"BASE_DIR = {BASE_DIR}")
    log(f"DB_FILE (exists?) = {DB_FILE}  /  {DB_FILE.exists()}")
    log(f"DB_PATH = {DB_PATH}")
    log(f"------------------------")


_debug_db_paths()

# ---------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------
def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")

def initialize_database():
    """
    Create database schema if tables do not exist, using models.sql.
    This only runs if pos.db exists but required tables are missing.
    """
    try:
        # Required tables
        required = {"items", "components", "recipes", "inventory", "orders", "order_lines"}
        
        with connect() as c:
            existing = {row[0] for row in c.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()}

            missing = required - existing
            if not missing:
                log(f"DB Init: All required tables already present.")
                return

            log(f"DB Init: Missing tables detected: {missing}")

            # Load models.sql from backend
            model_path = BASE_DIR / "backend" / "models.sql"
            if not model_path.exists():
                log(f"DB Init ERROR: models.sql not found at {model_path}")
                return

            sql_text = model_path.read_text(encoding="utf-8")
            c.executescript(sql_text)
            c.commit()

            log(f"DB Init: models.sql executed successfully — schema created.")

    except Exception as e:
        log(f"DB Init FAILED: {e}")

def make_safety_db_backup() -> str:
    """
    Copy the current database before a restore.
    Returns the path to the safety backup.
    """
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = _safety_backups_dir()

    src = DB_PATH
    filename = f"pos_safety_backup_{stamp}.db"
    dest = os.path.join(backup_dir, filename)

    shutil.copy2(src, dest)

    return dest

# ---------------------------------------------------------------------
# Migrations (run on app start)
# ---------------------------------------------------------------------
def _migrate_receipt_seq():
    """Ensure orders table has receipt_seq and receipt_date."""
    with connect() as c:
        cur = c.execute("PRAGMA table_info(orders)")
        cols = {row[1] for row in cur.fetchall()}
        if "receipt_seq" not in cols:
            c.execute("ALTER TABLE orders ADD COLUMN receipt_seq INTEGER")
        if "receipt_date" not in cols:
            c.execute("ALTER TABLE orders ADD COLUMN receipt_date TEXT")
        c.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS ux_orders_receipt
            ON orders(receipt_date, receipt_seq)
        """)
        c.commit()
    log(f"Migration: receipt sequence verified.")
        
def _migrate_orders_table():
    """Ensure the orders table exists."""
    with connect() as c:
        c.execute("""
        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ts TEXT NOT NULL,
            cashier TEXT NOT NULL,
            method TEXT NOT NULL,
            discount REAL NOT NULL DEFAULT 0,
            tax REAL NOT NULL DEFAULT 0,
            subtotal REAL NOT NULL DEFAULT 0,
            total REAL NOT NULL DEFAULT 0,
            cash_given REAL DEFAULT 0,
            change_given REAL DEFAULT 0,
            receipt_date TEXT,
            receipt_seq INTEGER,
            shift_id INTEGER,  -- REQUIRED so FK works
            FOREIGN KEY (shift_id) REFERENCES shifts(id)
            );
        """)
        c.commit()
    log(f"Migration: orders table verified.")
    
def _migrate_shifts_table():
    """Ensure the shifts table exists."""
    with connect() as c:
        c.execute("""
        CREATE TABLE IF NOT EXISTS shifts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ts_start TEXT NOT NULL,
            ts_end TEXT,
            cashier TEXT NOT NULL,
            opening_float REAL NOT NULL DEFAULT 0,
            closing_amount REAL,
            over_short REAL,
            is_active INTEGER NOT NULL DEFAULT 1
        );
        """)
        c.commit()
    log(f"Migration: shifts table verified.")
        
def _migrate_order_cash_columns():
    """Ensure cash tracking columns exist."""
    with connect() as c:
        # shifts
        cols = [r["name"] for r in c.execute("PRAGMA table_info(shifts)").fetchall()]
        if "cash_sales" not in cols:
            c.execute("ALTER TABLE shifts ADD COLUMN cash_sales REAL DEFAULT 0")
        if "cash_given" not in cols:
            c.execute("ALTER TABLE shifts ADD COLUMN cash_given REAL DEFAULT 0")
        if "change_given" not in cols:
            c.execute("ALTER TABLE shifts ADD COLUMN change_given REAL DEFAULT 0")

        # orders
        cols = [r["name"] for r in c.execute("PRAGMA table_info(orders)").fetchall()]
        if "cash_given" not in cols:
            c.execute("ALTER TABLE orders ADD COLUMN cash_given REAL DEFAULT 0")
        if "change_given" not in cols:
            c.execute("ALTER TABLE orders ADD COLUMN change_given REAL DEFAULT 0")

        c.commit()
    log(f"Migration: cash tracking columns verified.")
        
def _migrate_shift_last_sale():
    """Ensure last_sale_at exists."""
    with connect() as c:
        cols = [r["name"] for r in c.execute("PRAGMA table_info(shifts)").fetchall()]
        if "last_sale_at" not in cols:
            c.execute("ALTER TABLE shifts ADD COLUMN last_sale_at TEXT")
        c.commit()
    log(f"Migration: last_sale_at verified.")
    
def _migrate_discount_type():
    with connect() as c:
        cols = [r["name"] for r in c.execute("PRAGMA table_info(orders)").fetchall()]
        if "discount_type" not in cols:
            c.execute("ALTER TABLE orders ADD COLUMN discount_type TEXT")
            c.commit()
            log(f"Migration: discount_type column added.")
            
def _migrate_settings_tables():
    with connect() as c:
        c.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            )
        """)

        c.execute("""
            CREATE TABLE IF NOT EXISTS admin_pins (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pin TEXT NOT NULL,
                label TEXT,
                active INTEGER NOT NULL DEFAULT 1
            )
        """)

        for key in [
            "require_pin_discounts",
            "require_pin_admin_access",
            "require_pin_exit_program",
            "tax_rate"
        ]:
            c.execute("""
                INSERT OR IGNORE INTO settings(key, value)
                VALUES (?, '0')
            """, (key,))
            
        c.commit()

def _migrate_sales_table():
    """Create sales/promotions table and ensure BOGO fields exist."""
    with connect() as c:
        c.execute("""
            CREATE TABLE IF NOT EXISTS sales (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                sale_type TEXT NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                item_id INTEGER,
                category TEXT,
                buy_qty INTEGER NOT NULL DEFAULT 1,
                free_qty INTEGER NOT NULL DEFAULT 1,
                auto_apply INTEGER NOT NULL DEFAULT 0,
                active INTEGER NOT NULL DEFAULT 1,
                starts_on TEXT,
                ends_on TEXT,
                FOREIGN KEY (item_id) REFERENCES items(id)
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS sale_requirements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sale_id INTEGER NOT NULL,
                requirement_type TEXT NOT NULL DEFAULT 'CATEGORY',
                item_id INTEGER,
                category TEXT,
                qty INTEGER NOT NULL DEFAULT 1,
                FOREIGN KEY (sale_id) REFERENCES sales(id) ON DELETE CASCADE,
                FOREIGN KEY (item_id) REFERENCES items(id)
            )
        """)

        cols = [r["name"] for r in c.execute("PRAGMA table_info(sales)").fetchall()]

        if "category" not in cols:
            c.execute("ALTER TABLE sales ADD COLUMN category TEXT")

        if "buy_qty" not in cols:
            c.execute("ALTER TABLE sales ADD COLUMN buy_qty INTEGER NOT NULL DEFAULT 1")

        if "free_qty" not in cols:
            c.execute("ALTER TABLE sales ADD COLUMN free_qty INTEGER NOT NULL DEFAULT 1")
        
        if "auto_apply" not in cols:
            c.execute("ALTER TABLE sales ADD COLUMN auto_apply INTEGER NOT NULL DEFAULT 0")

        c.commit()

    log("Migration: sales table verified.")

def _migrate_order_discounts_table():
    """Track detailed discount records per order."""
    with connect() as c:
        c.execute("""
            CREATE TABLE IF NOT EXISTS order_discounts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id INTEGER NOT NULL,
                sale_id INTEGER,
                sale_name TEXT,
                sale_type TEXT,
                source TEXT NOT NULL, -- AUTO or MANUAL
                amount REAL NOT NULL DEFAULT 0,
                authorized_pin_id INTEGER,
                authorized_pin_label TEXT,
                created_at TEXT NOT NULL DEFAULT (datetime('now', '-6 hours')),
                FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE,
                FOREIGN KEY (sale_id) REFERENCES sales(id),
                FOREIGN KEY (authorized_pin_id) REFERENCES admin_pins(id)
            )
        """)
        c.commit()

    log("Migration: order_discounts table verified.")


def _migrate_order_line_modifiers_table():
    """Store POS modifier details such as flavor pumps for receipt/report display."""
    with connect() as c:
        c.execute("""
            CREATE TABLE IF NOT EXISTS order_line_modifiers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_line_id INTEGER NOT NULL,
                modifier_type TEXT NOT NULL DEFAULT 'flavor',
                name TEXT NOT NULL,
                qty INTEGER NOT NULL DEFAULT 1,
                unit_price REAL NOT NULL DEFAULT 0,
                line_total REAL NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL DEFAULT (datetime('now', '-6 hours')),
                FOREIGN KEY (order_line_id) REFERENCES order_lines(id) ON DELETE CASCADE
            )
        """)
        c.commit()

    log("Migration: order_line_modifiers table verified.")


def _migrate_till_counts_table():
    """Store opening and closing denomination counts for each shift."""
    with connect() as c:
        c.execute("""
            CREATE TABLE IF NOT EXISTS till_counts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                shift_id INTEGER NOT NULL,
                count_type TEXT NOT NULL,
                hundreds INTEGER NOT NULL DEFAULT 0,
                fifties INTEGER NOT NULL DEFAULT 0,
                twenties INTEGER NOT NULL DEFAULT 0,
                tens INTEGER NOT NULL DEFAULT 0,
                fives INTEGER NOT NULL DEFAULT 0,
                ones INTEGER NOT NULL DEFAULT 0,
                quarters INTEGER NOT NULL DEFAULT 0,
                dimes INTEGER NOT NULL DEFAULT 0,
                nickels INTEGER NOT NULL DEFAULT 0,
                pennies INTEGER NOT NULL DEFAULT 0,
                total REAL NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL DEFAULT (datetime('now', '-6 hours')),
                FOREIGN KEY (shift_id) REFERENCES shifts(id)
            )
        """)
        c.commit()

    log("Migration: till_counts table verified.")


def _migrate_item_pos_group_fields():
    """
    Add POS grouping fields so multiple item SKUs can appear as one drink/item
    button with size choices.

    category = broad POS category such as Drinks, Food, Specials.
    display_group = customer-facing grouped button name, e.g. Coffee.
    size_label = option label inside modal, e.g. 8oz or 12oz.
    allow_flavors = future toggle for modifier/flavor UI.
    pos_sort_order = optional manual sorting.
    """
    with connect() as c:
        cols = [r["name"] for r in c.execute("PRAGMA table_info(items)").fetchall()]

        if "display_group" not in cols:
            c.execute("ALTER TABLE items ADD COLUMN display_group TEXT")

        if "size_label" not in cols:
            c.execute("ALTER TABLE items ADD COLUMN size_label TEXT")

        if "allow_flavors" not in cols:
            c.execute("ALTER TABLE items ADD COLUMN allow_flavors INTEGER NOT NULL DEFAULT 0")

        if "pos_sort_order" not in cols:
            c.execute("ALTER TABLE items ADD COLUMN pos_sort_order INTEGER NOT NULL DEFAULT 0")

        # Sensible backfill for existing items:
        # "Coffee 8oz" -> display_group "Coffee", size_label "8oz"
        # "Coffee 12oz" -> display_group "Coffee", size_label "12oz"
        # Anything else keeps its full name as the display_group.
        rows = c.execute("""
            SELECT id, name, display_group, size_label
            FROM items
            WHERE active = 1
        """).fetchall()

        for r in rows:
            name = (r["name"] or "").strip()
            display_group = (r["display_group"] or "").strip()
            size_label = (r["size_label"] or "").strip()

            if not display_group or not size_label:
                parts = name.split()
                guessed_size = ""
                guessed_group = name

                if parts:
                    last = parts[-1].lower()
                    if last.endswith("oz") or last in {"small", "medium", "large"}:
                        guessed_size = parts[-1]
                        guessed_group = " ".join(parts[:-1]).strip() or name

                c.execute("""
                    UPDATE items
                    SET display_group = COALESCE(NULLIF(display_group, ''), ?),
                        size_label = COALESCE(NULLIF(size_label, ''), ?)
                    WHERE id = ?
                """, (
                    guessed_group,
                    guessed_size,
                    r["id"]
                ))

        c.commit()

    log("Migration: item POS grouping fields verified.")


def _migrate_flavors_table():
    """Create/manage POS flavor pump options. These are POS modifiers, not recipe-tracked inventory."""
    with connect() as c:
        c.execute("""
            CREATE TABLE IF NOT EXISTS flavors (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                price_per_pump REAL NOT NULL DEFAULT 0.25,
                active INTEGER NOT NULL DEFAULT 1,
                sort_order INTEGER NOT NULL DEFAULT 0
            )
        """)

        # Seed a practical starter list only when the table is empty.
        existing = c.execute("SELECT COUNT(*) AS n FROM flavors").fetchone()["n"]
        if int(existing or 0) == 0:
            starter_flavors = [
                ("Vanilla", 0.25, 1, 10),
                ("Caramel", 0.25, 1, 20),
                ("Hazelnut", 0.25, 1, 30),
                ("Mocha", 0.25, 1, 40),
                ("Peppermint", 0.25, 1, 50),
                ("Sugar Free Vanilla", 0.25, 1, 60),
            ]
            c.executemany("""
                INSERT OR IGNORE INTO flavors(name, price_per_pump, active, sort_order)
                VALUES (?, ?, ?, ?)
            """, starter_flavors)

        c.commit()

    log("Migration: flavors table verified.")


def list_flavors(include_inactive: bool = False) -> list[dict]:
    """Return POS flavor pump options."""
    with connect() as c:
        where = "" if include_inactive else "WHERE active = 1"
        rows = c.execute(f"""
            SELECT id, name, price_per_pump, active, sort_order
            FROM flavors
            {where}
            ORDER BY sort_order, name
        """).fetchall()
        return [dict(r) for r in rows]


def add_flavor(data: dict):
    """Create or update a flavor by name."""
    name = (data.get("name") or "").strip()
    if not name:
        raise ValueError("Flavor name is required.")

    price_per_pump = float(data.get("price_per_pump", 0.25) or 0.25)
    active = int(data.get("active", 1) or 0)
    sort_order = int(data.get("sort_order", 0) or 0)

    with connect() as c:
        c.execute("""
            INSERT INTO flavors(name, price_per_pump, active, sort_order)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(name) DO UPDATE SET
                price_per_pump=excluded.price_per_pump,
                active=excluded.active,
                sort_order=excluded.sort_order
        """, (name, price_per_pump, active, sort_order))
        c.commit()


def update_flavor(data: dict):
    """Update an existing flavor."""
    with connect() as c:
        c.execute("""
            UPDATE flavors
            SET name=?, price_per_pump=?, active=?, sort_order=?
            WHERE id=?
        """, (
            (data.get("name") or "").strip(),
            float(data.get("price_per_pump", 0.25) or 0.25),
            int(data.get("active", 1) or 0),
            int(data.get("sort_order", 0) or 0),
            int(data.get("id"))
        ))
        c.commit()


def delete_flavor(flavor_id: int):
    """Delete a flavor option."""
    with connect() as c:
        c.execute("DELETE FROM flavors WHERE id=?", (int(flavor_id),))
        c.commit()


_TILL_VALUES = {
    "hundreds": 100.00,
    "fifties": 50.00,
    "twenties": 20.00,
    "tens": 10.00,
    "fives": 5.00,
    "ones": 1.00,
    "quarters": 0.25,
    "dimes": 0.10,
    "nickels": 0.05,
    "pennies": 0.01,
}


def _normalize_till_count(till_count: dict | None) -> dict:
    till_count = till_count or {}
    normalized = {}

    for key in _TILL_VALUES:
        try:
            normalized[key] = max(0, int(till_count.get(key, 0) or 0))
        except Exception:
            normalized[key] = 0

    return normalized


def calculate_till_total(till_count: dict | None) -> float:
    counts = _normalize_till_count(till_count)
    total = sum(counts[key] * value for key, value in _TILL_VALUES.items())
    return round(total, 2)


def _save_till_count(conn, shift_id: int, count_type: str, till_count: dict | None) -> float:
    counts = _normalize_till_count(till_count)
    total = calculate_till_total(counts)

    conn.execute("""
        INSERT INTO till_counts (
            shift_id,
            count_type,
            hundreds,
            fifties,
            twenties,
            tens,
            fives,
            ones,
            quarters,
            dimes,
            nickels,
            pennies,
            total
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        shift_id,
        count_type,
        counts["hundreds"],
        counts["fifties"],
        counts["twenties"],
        counts["tens"],
        counts["fives"],
        counts["ones"],
        counts["quarters"],
        counts["dimes"],
        counts["nickels"],
        counts["pennies"],
        total
    ))

    return total

    
# ---------------------------------------------------------------------`
# Shifts (daily cashier sessions)
# ---------------------------------------------------------------------`
def start_shift(cashier: str, opening_float: float = 0, till_count: dict | None = None):
    """
    Start a shift.

    If till_count is provided, opening_float is calculated from the denomination
    count and a till_counts row is saved.
    """
    with connect() as c:
        active = c.execute("SELECT id FROM shifts WHERE is_active=1").fetchone()
        if active:
            raise ValueError("Shift already active.")

        opening_total = calculate_till_total(till_count) if till_count is not None else round(float(opening_float or 0), 2)

        cur = c.execute("""
            INSERT INTO shifts(ts_start, cashier, opening_float, is_active)
            VALUES(datetime('now', '-6 hours'), ?, ?, 1)
        """, (cashier, opening_total))

        shift_id = cur.lastrowid

        if till_count is not None:
            _save_till_count(c, shift_id, "opening", till_count)

        c.commit()

    log(f"Shift started for {cashier}")
    return {
        "ok": True,
        "shift_id": shift_id,
        "cashier": cashier,
        "opening_float": opening_total
    }



def get_active_shift():
    with connect() as c:
        row = c.execute("SELECT * FROM shifts WHERE is_active=1").fetchone()
        return dict(row) if row else None


def get_cash_summary(ts_start: str | None = None) -> dict:
    with connect() as c:
        if ts_start is None:
            shift = c.execute("SELECT * FROM shifts WHERE is_active=1").fetchone()
            if not shift:
                return {"opening": 0, "cash_sales": 0, "net_cash": 0, "expected": 0}
            opening = float(shift["opening_float"] or 0)
            ts_start = shift["ts_start"]
        else:
            sh = c.execute("SELECT opening_float FROM shifts WHERE ts_start=? LIMIT 1", (ts_start,)).fetchone()
            opening = float(sh["opening_float"] or 0) if sh else 0

        row = c.execute("""
            SELECT
                COALESCE(SUM(total), 0) AS cash_sales,
                COALESCE(SUM(cash_given - change_given), 0) AS net_cash
            FROM orders
            WHERE method='Cash' AND ts >= ?
        """, (ts_start,)).fetchone()

        cash_sales = float(row["cash_sales"] or 0)
        net_cash   = float(row["net_cash"] or 0)
        expected   = round(opening + net_cash, 2)

        return {
            "opening": round(opening, 2),
            "cash_sales": round(cash_sales, 2),
            "net_cash": round(net_cash, 2),
            "expected": expected
        }

# ---------------------------------------------------------------------
# Close shift
# ---------------------------------------------------------------------
def close_shift(actual_cash: float | None = None, till_count: dict | None = None):
    """
    Close the current active shift.
    Calculates expected cash, computes over/short, and finalizes the record.

    If till_count is provided, actual_cash is calculated from the denomination count
    and a till_counts row is saved.
    """
    with connect() as c:
        shift = c.execute("SELECT * FROM shifts WHERE is_active=1").fetchone()
        if not shift:
            raise ValueError("No active shift found.")

        shift_id = shift["id"]
        ts_start = shift["ts_start"]
        opening = float(shift["opening_float"] or 0)

        if till_count is not None:
            actual_cash = _save_till_count(c, shift_id, "closing", till_count)
        else:
            actual_cash = float(actual_cash or 0)

        # --- Calculate net cash movement since shift start ---
        row = c.execute("""
            SELECT 
                COALESCE(SUM(cash_given - change_given), 0) AS net_cash
            FROM orders
            WHERE method='Cash' AND ts >= ?
        """, (ts_start,)).fetchone()

        net_cash = float(row["net_cash"] or 0)
        expected = round(opening + net_cash, 2)
        diff = round(float(actual_cash) - expected, 2)

        # --- Update shift record ---
        c.execute("""
            UPDATE shifts
            SET ts_end = datetime('now', '-6 hours'),
                closing_amount = ?,
                over_short = ?,
                is_active = 0
            WHERE id = ?
        """, (actual_cash, diff, shift_id))
        c.commit()

        return {
            "cashier": shift["cashier"],
            "expected": expected,
            "actual": round(actual_cash, 2),
            "over_short": diff
        }

# ---------------------------------------------------------------------
# Close stale shift and log expected till
# ---------------------------------------------------------------------
def _auto_close_stale_shifts():
    """Auto-close any left-open shifts from previous days."""
    with connect() as c:
        rows = c.execute("""
            SELECT id, cashier, ts_start, opening_float
            FROM shifts
            WHERE is_active = 1
        """).fetchall()

        if not rows:
            log("No stale shifts found.")
            return

        closed = 0

        # Use same local-time offset as the rest of the POS
        today = c.execute("SELECT date('now', '-6 hours')").fetchone()[0]

        for r in rows:
            # Works for timestamps like '2025-10-11 14:22:00'
            start_date = (r["ts_start"] or "")[:10]

            if start_date != today:
                net_cash = c.execute("""
                    SELECT COALESCE(SUM(cash_given - change_given), 0)
                    FROM orders
                    WHERE method='Cash' AND ts >= ?
                """, (r["ts_start"],)).fetchone()[0]

                expected = float(r["opening_float"] or 0) + float(net_cash or 0)

                log(
                    f"Auto-close stale shift: "
                    f"cashier={r['cashier']} "
                    f"start={r['ts_start']} "
                    f"expected_till={expected}"
                )

                c.execute("""
                    UPDATE shifts
                    SET is_active = 0,
                        ts_end = datetime('now', '-6 hours'),
                        closing_amount = ?,
                        over_short = NULL
                    WHERE id = ?
                """, (expected, r["id"]))

                closed += 1

        c.commit()

        if closed:
            log(f"Auto-closed {closed} stale shift(s).")
        else:
            log("All active shifts are from today; no auto-close needed.")
            
# ---------------------------------------------------------------------
# Items (what appears on the POS)
# ---------------------------------------------------------------------
def list_items():
    """Return all active sellable items with calculated recipe cost."""
    with connect() as c:
        rows = c.execute("""
            SELECT
                i.id,
                i.sku,
                i.name,
                i.category,
                i.price,
                i.active,
                COALESCE(NULLIF(i.display_group, ''), i.name) AS display_group,
                COALESCE(i.size_label, '') AS size_label,
                COALESCE(i.allow_flavors, 0) AS allow_flavors,
                COALESCE(i.pos_sort_order, 0) AS pos_sort_order,

                ROUND(
                    COALESCE(SUM(
                        r.qty_per_item *
                        CASE
                            WHEN inv.units_per_case > 0
                            THEN inv.case_cost / inv.units_per_case
                            ELSE 0
                        END
                    ), 0),
                2) AS item_cost

            FROM items i
            LEFT JOIN recipes r ON r.item_id = i.id
            LEFT JOIN components comp ON comp.id = r.component_id
            LEFT JOIN inventory inv ON inv.id = comp.inventory_id

            WHERE i.active = 1

            GROUP BY
                i.id,
                i.sku,
                i.name,
                i.category,
                i.price,
                i.active,
                i.display_group,
                i.size_label,
                i.allow_flavors,
                i.pos_sort_order

            ORDER BY i.category, pos_sort_order, display_group, size_label, i.name
        """).fetchall()

        items = []

        for r in rows:
            item = dict(r)

            price = float(item.get("price") or 0)
            cost = float(item.get("item_cost") or 0)

            item["gross_profit"] = round(price - cost, 2)

            if price > 0:
                item["margin_percent"] = round(((price - cost) / price) * 100, 1)
            else:
                item["margin_percent"] = 0

            items.append(item)

        return items

def get_item_by_sku(conn, sku:str):
    row = conn.execute("SELECT id, price FROM items WHERE sku=? OR name=?", (sku, sku)).fetchone()
    return dict(row) if row else None

def add_component(data):
    """Add or update a component linked to inventory by inventory_id."""
    inventory_id = int(data.get("inventory_id") or 0)
    display_name = (data.get("display_name") or "").strip() or None
    pos_track_sellout = int(data.get("pos_track_sellout", 0))

    if inventory_id <= 0:
        raise ValueError("inventory_id is required.")

    with connect() as c:
        inv = c.execute("""
            SELECT id, name
            FROM inventory
            WHERE id = ?
            LIMIT 1
        """, (inventory_id,)).fetchone()

        if not inv:
            raise ValueError("Inventory item not found.")

        inventory_name = inv["name"]

        existing = c.execute("""
            SELECT id
            FROM components
            WHERE inventory_id = ?
            LIMIT 1
        """, (inventory_id,)).fetchone()

        if existing:
            c.execute("""
                UPDATE components
                SET name = ?,
                    display_name = COALESCE(?, display_name),
                    pos_track_sellout = ?
                WHERE id = ?
            """, (
                inventory_name,
                display_name,
                pos_track_sellout,
                existing["id"]
            ))
        else:
            c.execute("""
                INSERT INTO components (
                    name,
                    display_name,
                    inventory_id,
                    pos_track_sellout
                )
                VALUES (?, ?, ?, ?)
            """, (
                inventory_name,
                display_name,
                inventory_id,
                pos_track_sellout
            ))

        c.commit()

# ---------------------------------------------------------------------
# Availability / Stock checks
# ---------------------------------------------------------------------
def available_qty(item_id: int) -> int:
    """Return how many of this item can be sold before components run out."""
    with connect() as c:
        rows = c.execute("""
            SELECT inv.qty_on_hand AS onhand, r.qty_per_item AS per
            FROM recipes r
            JOIN components comp ON comp.id = r.component_id
            JOIN inventory inv ON inv.id = comp.inventory_id
            WHERE r.item_id=? AND comp.pos_track_sellout=1
        """, (item_id,)).fetchall()

        if not rows:
            return 10**9  # no limiting components → unlimited

        limits = []
        for r in rows:
            per = float(r["per"] or 0)
            if per <= 0:
                continue
            onhand = float(r["onhand"] or 0)
            limits.append(int(onhand // per))

        return min(limits) if limits else 10**9

def sold_out_items():
    """List of items whose limiting component = 0."""
    with connect() as c:
        items = c.execute("SELECT id, name FROM items WHERE active=1").fetchall()
        sold_out = []
        for it in items:
            if available_qty(it["id"]) <= 0:
                sold_out.append(dict(it))
        return sold_out
    

# ---------------------------------------------------------------------
# Orders & Sales
# ---------------------------------------------------------------------
def record_order(cashier:str, method:str, discount:float, lines:list, discount_type:str=None,
                 discount_details:list=None,
                 tax_rate:float=0.0, cash_given:float=0.0, change_given:float=0.0):
    """Record a sale, update inventory, and update shift cash totals."""
    with connect() as c:
        ts = c.execute("SELECT datetime('now', '-6 hours')").fetchone()[0]

        shift_row = c.execute("SELECT id FROM shifts WHERE is_active=1").fetchone()
        shift_id = shift_row["id"] if shift_row else None
        
        # --- Build cart ---
        cart = []
        for l in lines:
            sku = l.get("item_sku")
            it = get_item_by_sku(c, sku)
            if not it:
                raise ValueError(f"Unknown item: {sku}")

            base_price = float(it["price"])

            # POS modifiers such as flavor pumps may send a configured unit_price.
            # We accept higher configured prices, but never allow a client-side override
            # to reduce the item below its base price. Discounts should use discount logic.
            requested_unit_price = l.get("unit_price")
            if requested_unit_price is None:
                unit_price = base_price
            else:
                unit_price = max(base_price, round(float(requested_unit_price or base_price), 2))

            modifiers = []

            raw_modifiers = l.get("modifiers", []) or []

            # Current POS shape may be:
            #   modifiers: { flavors: [{name, pumps, price_per_pump}, ...] }
            # Older/intermediate shapes may be:
            #   modifiers: [{name, qty, unit_price}, ...]
            #   modifiers: ["Vanilla", "Caramel"]
            # Normalize all of them into one flat list of modifier dicts.
            if isinstance(raw_modifiers, dict):
                flattened = []

                for f in raw_modifiers.get("flavors", []) or []:
                    if isinstance(f, dict):
                        f = dict(f)
                        f.setdefault("type", "flavor")
                        flattened.append(f)
                    elif isinstance(f, str):
                        flattened.append({"type": "flavor", "name": f, "qty": 1, "unit_price": 0.25})

                # Keep support for any future modifier groups too.
                for key, value in raw_modifiers.items():
                    if key == "flavors":
                        continue
                    if isinstance(value, list):
                        for m in value:
                            if isinstance(m, dict):
                                m = dict(m)
                                m.setdefault("type", key.rstrip("s") or "modifier")
                                flattened.append(m)
                            elif isinstance(m, str):
                                flattened.append({"type": key.rstrip("s") or "modifier", "name": m, "qty": 1, "unit_price": 0})

                raw_modifiers = flattened

            for m in raw_modifiers:
                if isinstance(m, str):
                    name = m.strip()
                    qty = 1
                    mod_unit_price = 0.25
                    mod_type = "flavor"
                elif isinstance(m, dict):
                    name = (m.get("name") or m.get("label") or "").strip()
                    qty = int(m.get("qty", m.get("pumps", m.get("count", 0))) or 0)
                    mod_unit_price = round(float(
                        m.get("unit_price", m.get("price_per_pump", m.get("price", 0.25 if (m.get("type") or "flavor") == "flavor" else 0))) or 0
                    ), 2)
                    mod_type = (m.get("type") or m.get("modifier_type") or "flavor").strip() or "flavor"
                else:
                    continue

                if not name or qty <= 0:
                    continue

                modifiers.append({
                    "type": mod_type,
                    "name": name,
                    "qty": qty,
                    "unit_price": mod_unit_price,
                    "line_total": round(qty * mod_unit_price, 2)
                })

            cart.append({
                "item_id": it["id"],
                "qty": float(l["qty"]),
                "unit_price": unit_price,
                "modifiers": modifiers
            })

        subtotal = sum(x["qty"] * x["unit_price"] for x in cart)
        taxable  = max(0.0, subtotal - float(discount or 0))
        tax      = round(taxable * tax_rate, 2)
        total    = taxable + tax
        rdate    = _today_date(ts)
        rseq     = _next_receipt_seq(rdate)

        # --- Create order ---
        cur = c.execute("""
            INSERT INTO orders(
                ts, cashier, method,
                discount, discount_type,
                tax, subtotal, total,
                cash_given, change_given,
                receipt_date, receipt_seq,
                shift_id
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            ts,
            cashier,
            method,
            discount,
            discount_type,
            tax,
            subtotal,
            total,
            cash_given,
            change_given,
            rdate,
            rseq,
            shift_id
        ))

        order_id = cur.lastrowid
        
        # --- Record discount details ---
        for d in discount_details or []:
            amount = float(d.get("amount") or 0)

            if amount <= 0:
                continue

            c.execute("""
                INSERT INTO order_discounts (
                    order_id,
                    sale_id,
                    sale_name,
                    sale_type,
                    source,
                    amount,
                    authorized_pin_id,
                    authorized_pin_label
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                order_id,
                int(d["sale_id"]) if d.get("sale_id") else None,
                d.get("sale_name"),
                d.get("sale_type"),
                d.get("source", "MANUAL"),
                amount,
                int(d["authorized_pin_id"]) if d.get("authorized_pin_id") else None,
                d.get("authorized_pin_label")
            ))

        # --- Order lines ---
        for x in cart:
            cur_line = c.execute("""
                INSERT INTO order_lines(order_id, item_id, qty, unit_price)
                VALUES (?, ?, ?, ?)
            """, (order_id, x["item_id"], x["qty"], x["unit_price"]))

            order_line_id = cur_line.lastrowid

            for m in x.get("modifiers", []) or []:
                c.execute("""
                    INSERT INTO order_line_modifiers (
                        order_line_id,
                        modifier_type,
                        name,
                        qty,
                        unit_price,
                        line_total
                    )
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (
                    order_line_id,
                    m.get("type", "flavor"),
                    m.get("name", ""),
                    int(m.get("qty", 0) or 0),
                    float(m.get("unit_price", 0) or 0),
                    float(m.get("line_total", 0) or 0)
                ))

        # --- Update inventory for each component ---
        rows = c.execute("""
            SELECT ol.item_id,
                   SUM(ol.qty) AS qty_sold,
                   inv.id      AS inv_id,
                   inv.name    AS inv_name,
                   r.qty_per_item AS per_item
            FROM order_lines ol
            JOIN recipes r ON r.item_id = ol.item_id
            JOIN components c2 ON c2.id = r.component_id
            JOIN inventory inv ON inv.id = c2.inventory_id
            WHERE ol.order_id = ?
            GROUP BY ol.item_id, inv.id, r.qty_per_item
        """, (order_id,)).fetchall()

        touched_items = set()
        for r in rows:
            touched_items.add(r["item_id"])
            used_qty = float(r["qty_sold"] or 0) * float(r["per_item"] or 0)
            if used_qty <= 0:
                continue
            c.execute("""
                UPDATE inventory
                SET qty_on_hand = MAX(qty_on_hand - ?, 0)
                WHERE id = ?
            """, (used_qty, r["inv_id"]))

        # --- Fallback for items with no recipe link ---
        leftovers = c.execute("""
            SELECT item_id, SUM(qty) AS qty
            FROM order_lines
            WHERE order_id = ?
            GROUP BY item_id
        """, (order_id,)).fetchall()

        for lf in leftovers:
            if lf["item_id"] in touched_items:
                continue
            qty = float(lf["qty"] or 0)
            if qty <= 0:
                continue
            lim = c.execute("""
                SELECT inv.id
                FROM recipes r
                JOIN components c3 ON c3.id = r.component_id
                JOIN inventory inv ON inv.id = c3.inventory_id
                WHERE r.item_id=? AND c3.pos_track_sellout=1
                ORDER BY c3.id LIMIT 1
            """, (lf["item_id"],)).fetchone()
            if lim:
                c.execute("""
                    UPDATE inventory
                    SET qty_on_hand = MAX(qty_on_hand - ?, 0)
                    WHERE id = ?
                """, (qty, lim["id"]))

        # --- Commit order + inventory changes ---
        c.commit()

        # --- Update shift totals if this was a cash order ---
        if method.lower() == "cash":
            try:
                c.execute("""
                    UPDATE shifts
                    SET
                        cash_sales   = COALESCE(cash_sales,0) + ?,
                        cash_given   = COALESCE(cash_given,0) + ?,
                        change_given = COALESCE(change_given,0) + ?,
                        last_sale_at = datetime('now', '-6 hours')
                    WHERE is_active = 1
                """, (total, cash_given, change_given))
                c.commit()
            except Exception as e:
                log(f"Could not update active shift totals: {e}")

    return {
        "order_id": order_id,
        "receipt_no": rseq,
        "date": rdate, 
        "total": total,
        "subtotal": subtotal,
        "tax": tax,
        "cash_given": cash_given,
        "change_given": change_given
    }

# ---------------------------------------------------------------------
# Sales / Promotions
# ---------------------------------------------------------------------

def list_sales(include_inactive: bool = False):
    """Return sales/promotions."""
    with connect() as c:
        where = "" if include_inactive else "WHERE s.active = 1"
        rows = c.execute(f"""
            SELECT
                s.id,
                s.name,
                s.sale_type,
                s.amount,
                s.item_id,
                s.category,
                s.buy_qty,
                s.free_qty,
                s.auto_apply,
                i.name AS item_name,
                s.active,
                s.starts_on,
                s.ends_on
            FROM sales s
            LEFT JOIN items i ON i.id = s.item_id
            {where}
            ORDER BY s.active DESC, s.name
        """).fetchall()

        sales = [dict(r) for r in rows]

        for s in sales:
            s["requirements"] = list_sale_requirements(s["id"])

        return sales


def list_active_sales():
    """Return only currently active sales."""
    today = datetime.now().date().isoformat()

    with connect() as c:
        rows = c.execute("""
            SELECT
                s.id,
                s.name,
                s.sale_type,
                s.amount,
                s.item_id,
                s.category,
                s.buy_qty,
                s.free_qty,
                s.auto_apply,
                i.name AS item_name,
                s.active,
                s.starts_on,
                s.ends_on
            FROM sales s
            LEFT JOIN items i ON i.id = s.item_id
            WHERE s.active = 1
              AND (s.starts_on IS NULL OR s.starts_on = '' OR s.starts_on <= ?)
              AND (s.ends_on IS NULL OR s.ends_on = '' OR s.ends_on >= ?)
            ORDER BY s.name
        """, (today, today)).fetchall()

        sales = [dict(r) for r in rows]

        for s in sales:
            s["requirements"] = list_sale_requirements(s["id"])

        return sales

def list_sale_requirements(sale_id: int):
    with connect() as c:
        rows = c.execute("""
            SELECT
                sr.id,
                sr.sale_id,
                sr.requirement_type,
                sr.item_id,
                i.name AS item_name,
                sr.category,
                sr.qty
            FROM sale_requirements sr
            LEFT JOIN items i ON i.id = sr.item_id
            WHERE sr.sale_id = ?
            ORDER BY sr.id
        """, (sale_id,)).fetchall()

        return [dict(r) for r in rows]
    
def replace_sale_requirements(sale_id: int, requirements: list):
    with connect() as c:
        c.execute("DELETE FROM sale_requirements WHERE sale_id=?", (sale_id,))

        for req in requirements or []:
            requirement_type = req.get("requirement_type", "CATEGORY")
            qty = int(req.get("qty") or 1)

            if qty < 1:
                qty = 1

            c.execute("""
                INSERT INTO sale_requirements (
                    sale_id,
                    requirement_type,
                    item_id,
                    category,
                    qty
                )
                VALUES (?, ?, ?, ?, ?)
            """, (
                sale_id,
                requirement_type,
                int(req["item_id"]) if req.get("item_id") else None,
                (req.get("category") or "").strip() or None,
                qty
            ))

        c.commit()
    
def add_sale(data):
    """Create a sale/promotion."""
    sale_type = data.get("sale_type", "DOLLAR_OFF")

    buy_qty = int(data.get("buy_qty") or 1)
    free_qty = int(data.get("free_qty") or 1)

    if buy_qty < 1:
        buy_qty = 1

    if free_qty < 1:
        free_qty = 1

    with connect() as c:
        cur = c.execute("""
            INSERT INTO sales (
                name,
                sale_type,
                amount,
                item_id,
                category,
                buy_qty,
                free_qty,
                auto_apply,
                active,
                starts_on,
                ends_on
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            (data.get("name") or "").strip(),
            sale_type,
            float(data.get("amount") or 0),
            int(data["item_id"]) if data.get("item_id") else None,
            (data.get("category") or "").strip() or None,
            buy_qty,
            free_qty,
            int(data.get("auto_apply", 0)),
            int(data.get("active", 1)),
            data.get("starts_on") or None,
            data.get("ends_on") or None
        ))

        sale_id = cur.lastrowid
        c.commit()

    if sale_type == "COMBO_PRICE":
        replace_sale_requirements(sale_id, data.get("requirements", []))
    
def update_sale(data):
    """Update an existing sale/promotion."""
    sale_type = data.get("sale_type", "DOLLAR_OFF")

    buy_qty = int(data.get("buy_qty") or 1)
    free_qty = int(data.get("free_qty") or 1)

    if buy_qty < 1:
        buy_qty = 1

    if free_qty < 1:
        free_qty = 1

    with connect() as c:
        c.execute("""
            UPDATE sales
            SET name=?,
                sale_type=?,
                amount=?,
                item_id=?,
                category=?,
                buy_qty=?,
                free_qty=?,
                auto_apply=?,
                active=?,
                starts_on=?,
                ends_on=?
            WHERE id=?
        """, (
            (data.get("name") or "").strip(),
            sale_type,
            float(data.get("amount") or 0),
            int(data["item_id"]) if data.get("item_id") else None,
            (data.get("category") or "").strip() or None,
            buy_qty,
            free_qty,
            int(data.get("auto_apply", 0)),
            int(data.get("active", 1)),
            data.get("starts_on") or None,
            data.get("ends_on") or None,
            int(data.get("id"))
        ))
        c.commit()
        
    if sale_type == "COMBO_PRICE":
        replace_sale_requirements(int(data.get("id")), data.get("requirements", []))
    else:
        replace_sale_requirements(int(data.get("id")), [])

def delete_sale(sale_id: int):
    """Delete a sale/promotion."""
    with connect() as c:
        c.execute("DELETE FROM sales WHERE id=?", (sale_id,))
        c.commit()
         
# ---------------------------------------------------------------------
# Admin helpers for adding new records
# ---------------------------------------------------------------------
def add_item(data):
    """Create or update a POS item safely (parentheses, special chars allowed)."""
    sku  = (data.get("sku") or "").strip()
    name = (data.get("name") or "").strip()
    cat  = (data.get("category") or "").strip()
    price = float(data.get("price") or 0)

    if not name:
        raise ValueError("Item name cannot be blank.")

    with connect() as c:
        # If SKU empty, auto-generate one from name
        if not sku:
            sku = name.replace(" ", "_").replace("(", "").replace(")", "").upper()[:12]

        display_group = (data.get("display_group") or "").strip() or name
        size_label = (data.get("size_label") or "").strip()
        allow_flavors = int(data.get("allow_flavors", 0) or 0)
        pos_sort_order = int(data.get("pos_sort_order", 0) or 0)

        c.execute("""
            INSERT INTO items (
                sku,
                name,
                category,
                price,
                active,
                display_group,
                size_label,
                allow_flavors,
                pos_sort_order
            )
            VALUES (?, ?, ?, ?, 1, ?, ?, ?, ?)
            ON CONFLICT(sku) DO UPDATE SET
              name=excluded.name,
              category=excluded.category,
              price=excluded.price,
              active=1,
              display_group=excluded.display_group,
              size_label=excluded.size_label,
              allow_flavors=excluded.allow_flavors,
              pos_sort_order=excluded.pos_sort_order
        """, (
            sku,
            name,
            cat,
            price,
            display_group,
            size_label,
            allow_flavors,
            pos_sort_order
        ))
        c.commit()
        
def update_item(data):
    """Edit an existing item (name, category, price)."""
    with connect() as c:
        cols = [r["name"] for r in c.execute("PRAGMA table_info(items)").fetchall()]

        if "display_group" in cols:
            c.execute("""
                UPDATE items
                SET name=?,
                    category=?,
                    price=?,
                    display_group=?,
                    size_label=?,
                    allow_flavors=?,
                    pos_sort_order=?
                WHERE id=?
            """, (
                data.get("name", "").strip(),
                data.get("category", "").strip(),
                float(data.get("price", 0)),
                (data.get("display_group") or "").strip() or data.get("name", "").strip(),
                (data.get("size_label") or "").strip(),
                int(data.get("allow_flavors", 0) or 0),
                int(data.get("pos_sort_order", 0) or 0),
                int(data.get("id"))
            ))
        else:
            c.execute("""
                UPDATE items
                SET name=?, category=?, price=?
                WHERE id=?
            """, (
                data.get("name", "").strip(),
                data.get("category", "").strip(),
                float(data.get("price", 0)),
                int(data.get("id"))
            ))
        c.commit()

def delete_item(item_id:int):
    """Delete an item and its linked recipes."""
    with connect() as c:
        c.execute("DELETE FROM recipes WHERE item_id=?", (item_id,))
        c.execute("DELETE FROM items WHERE id=?", (item_id,))
        c.commit()

# ---------------------------------------------------------------------
# COMPONENTS (Usage Tracking)
# ---------------------------------------------------------------------
def list_components():
    """Return components with stable inventory link, unit cost, nickname, and recipe usage details."""
    with connect() as c:
        rows = c.execute("""
            SELECT 
                comp.id,
                comp.name,
                comp.display_name,
                comp.inventory_id,
                inv.name AS inventory_name,

                COALESCE(
                    CASE 
                        WHEN inv.units_per_case > 0
                        THEN inv.case_cost / inv.units_per_case
                        ELSE 0
                    END,
                0) AS unit_cost,

                comp.pos_track_sellout,

                COUNT(DISTINCT r.item_id) AS used_in_count

            FROM components comp
            LEFT JOIN inventory inv ON inv.id = comp.inventory_id
            LEFT JOIN recipes r ON r.component_id = comp.id

            GROUP BY
                comp.id,
                comp.name,
                comp.display_name,
                comp.inventory_id,
                inv.name,
                inv.case_cost,
                inv.units_per_case,
                comp.pos_track_sellout

            ORDER BY COALESCE(comp.display_name, inv.name, comp.name), comp.name
        """).fetchall()

        comps = []

        for r in rows:
            comp = dict(r)

            inventory_name = comp.get("inventory_name") or comp.get("name") or ""
            nickname = comp.get("display_name") or ""

            if nickname and nickname != inventory_name:
                comp["label"] = f"{inventory_name} — {nickname}"
            else:
                comp["label"] = inventory_name

            used_rows = c.execute("""
                SELECT 
                    i.id AS item_id,
                    i.name AS item_name,
                    r.qty_per_item
                FROM recipes r
                JOIN items i ON i.id = r.item_id
                WHERE r.component_id = ?
                ORDER BY i.name
            """, (comp["id"],)).fetchall()

            comp["used_in"] = [dict(x) for x in used_rows]
            comps.append(comp)

        return comps

def update_component(data):
    """Update one component's nickname and sold-out tracking."""
    with connect() as c:
        c.execute("""
            UPDATE components
            SET display_name=?,
                pos_track_sellout=?
            WHERE id=?
        """, (
            (data.get("display_name") or "").strip() or None,
            int(data.get("pos_track_sellout", 0)),
            int(data.get("id"))
        ))
        c.commit()

def delete_component(comp_id:int):
    """Delete a component and any linked recipes."""
    with connect() as c:
        c.execute("DELETE FROM recipes WHERE component_id=?", (comp_id,))
        c.execute("DELETE FROM components WHERE id=?", (comp_id,))
        c.commit()

def batch_update_components(components: list):
    """Update multiple component nicknames and sold-out tracking at once."""
    with connect() as c:
        for comp in components:
            comp_id = int(comp.get("id"))
            display_name = (comp.get("display_name") or "").strip()
            track = int(comp.get("pos_track_sellout") or 0)

            c.execute("""
                UPDATE components
                SET display_name = ?,
                    pos_track_sellout = ?
                WHERE id = ?
            """, (
                display_name or None,
                track,
                comp_id
            ))

        c.commit()

def _migrate_component_display_name():
    """Add display_name/nickname field to components."""
    with connect() as c:
        cols = [r["name"] for r in c.execute("PRAGMA table_info(components)").fetchall()]

        if "display_name" not in cols:
            c.execute("ALTER TABLE components ADD COLUMN display_name TEXT")

        c.commit()

    log("Migration: component display_name verified.")
    
def delete_unlinked_blank_components():
    with connect() as c:
        c.execute("""
            DELETE FROM components
            WHERE inventory_id IS NULL
              AND (name IS NULL OR TRIM(name) = '')
              AND id NOT IN (
                  SELECT DISTINCT component_id
                  FROM recipes
              )
        """)
        c.commit()
    
def debug_bad_components():
    with connect() as c:
        rows = c.execute("""
            SELECT
                comp.id,
                comp.name,
                comp.display_name,
                comp.inventory_id,
                inv.name AS inventory_name,
                inv.case_cost,
                inv.units_per_case,
                CASE
                    WHEN inv.units_per_case > 0
                    THEN inv.case_cost / inv.units_per_case
                    ELSE 0
                END AS unit_cost
            FROM components comp
            LEFT JOIN inventory inv ON inv.id = comp.inventory_id
            ORDER BY comp.id
        """).fetchall()

        return [dict(r) for r in rows]

# ---------------------------------------------------------------------
# RECIPES
# ---------------------------------------------------------------------
def add_recipe(data):
    """
    Link an item to a component using stable IDs.

    Preferred:
      - item_id or item_key for the POS item
      - component_id for the component

    Legacy fallback:
      - component_name still works, but only as a backup.
    """
    with connect() as c:
        if data.get("item_id"):
            item = c.execute("""
                SELECT id
                FROM items
                WHERE id = ?
                LIMIT 1
            """, (int(data.get("item_id")),)).fetchone()
        else:
            item_key = data.get("item_key")
            item = c.execute("""
                SELECT id
                FROM items
                WHERE sku = ? OR name = ?
                LIMIT 1
            """, (item_key, item_key)).fetchone()

        if data.get("component_id"):
            comp = c.execute("""
                SELECT id
                FROM components
                WHERE id = ?
                LIMIT 1
            """, (int(data.get("component_id")),)).fetchone()
        else:
            component_name = (data.get("component_name") or "").strip()

            comp = c.execute("""
                SELECT comp.id
                FROM components comp
                LEFT JOIN inventory inv ON inv.id = comp.inventory_id
                WHERE inv.name = ?
                   OR comp.name = ?
                   OR comp.display_name = ?
                LIMIT 1
            """, (component_name, component_name, component_name)).fetchone()

        if not item:
            raise ValueError("Unknown item.")

        if not comp:
            raise ValueError("Unknown component.")

        qty = float(data.get("qty_per_item", 1) or 1)

        if qty <= 0:
            raise ValueError("qty_per_item must be greater than 0.")

        c.execute("""
            INSERT INTO recipes (item_id, component_id, qty_per_item)
            VALUES (?, ?, ?)
        """, (item["id"], comp["id"], qty))

        c.commit()

def update_recipe(data):
    """Update recipe quantity per item."""
    with connect() as c:
        c.execute("""
            UPDATE recipes
            SET qty_per_item=?
            WHERE id=?
        """, (float(data.get("qty_per_item", 1)), int(data.get("id"))))
        c.commit()

def delete_recipe(recipe_id:int):
    """Delete a recipe link."""
    with connect() as c:
        c.execute("DELETE FROM recipes WHERE id=?", (recipe_id,))
        c.commit()
        
def list_recipes():
    with connect() as c:
        rows = c.execute("""
            SELECT 
                r.id,
                i.name AS item_name,

                COALESCE(
                    NULLIF(c.display_name, ''),
                    inv.name,
                    c.name
                ) AS component_name,

                COALESCE(inv.name, c.name) AS component_inventory_name,
                c.display_name AS component_display_name,
                r.qty_per_item,

                ROUND(
                    CASE 
                        WHEN inv.units_per_case > 0 
                        THEN inv.case_cost / inv.units_per_case 
                        ELSE 0 
                    END,
                4) AS unit_cost,

                ROUND(
                    r.qty_per_item *
                    CASE 
                        WHEN inv.units_per_case > 0 
                        THEN inv.case_cost / inv.units_per_case 
                        ELSE 0 
                    END,
                4) AS line_cost

            FROM recipes r
            JOIN items i ON i.id = r.item_id
            JOIN components c ON c.id = r.component_id
            LEFT JOIN inventory inv ON inv.id = c.inventory_id

            ORDER BY i.name, COALESCE(NULLIF(c.display_name, ''), inv.name, c.name)
        """).fetchall()

        return [dict(r) for r in rows]
    
# ---------------------------------------------------------------------
# INVENTORY (Warehouse Stock)
# ---------------------------------------------------------------------
def list_inventory():
    """Return all inventory rows with calculated unit cost."""
    with connect() as c:
        rows = c.execute("""
            SELECT id,
                   name,
                   unit,
                   case_cost,
                   units_per_case,
                   CASE WHEN units_per_case > 0 
                        THEN ROUND(case_cost / units_per_case, 2)
                        ELSE 0 END AS unit_cost,
                   qty_on_hand,
                   reorder_point
            FROM inventory
            ORDER BY name
        """).fetchall()
        return [dict(r) for r in rows]

def add_inventory_item(data):
    """Add or update an inventory item."""
    with connect() as c:
        c.execute("""
            INSERT INTO inventory (name, unit, case_cost, units_per_case,
                                   qty_on_hand, reorder_point)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(name) DO UPDATE SET
              unit=excluded.unit,
              case_cost=excluded.case_cost,
              units_per_case=excluded.units_per_case,
              qty_on_hand=excluded.qty_on_hand,
              reorder_point=excluded.reorder_point
        """, (
            data.get("name", "").strip(),
            data.get("unit", "").strip(),
            float(data.get("case_cost", 0)),
            float(data.get("units_per_case", 0)),
            float(data.get("qty_on_hand", 0)),
            float(data.get("reorder_point", 0))
        ))
        c.commit()

def update_inventory_item(data):
    """Edit an existing inventory record."""
    with connect() as c:
        c.execute("""
            UPDATE inventory
            SET name=?,
                unit=?,
                case_cost=?,
                units_per_case=?,
                qty_on_hand=?,
                reorder_point=?
            WHERE id=?
        """, (
            data.get("name", "").strip(),
            data.get("unit", "").strip(),
            float(data.get("case_cost", 0)),
            float(data.get("units_per_case", 0)),
            float(data.get("qty_on_hand", 0)),
            float(data.get("reorder_point", 0)),
            int(data.get("id"))
        ))
        c.commit()

def delete_inventory_item(item_id:int):
    """Remove an inventory record."""
    with connect() as c:
        c.execute("DELETE FROM inventory WHERE id=?", (item_id,))
        c.commit()

def _migrate_component_inventory_id():
    """Add stable inventory_id link to components and backfill existing rows by name."""
    with connect() as c:
        cols = [r["name"] for r in c.execute("PRAGMA table_info(components)").fetchall()]

        if "inventory_id" not in cols:
            c.execute("ALTER TABLE components ADD COLUMN inventory_id INTEGER")

        # Backfill existing components by matching old component name to inventory name.
        c.execute("""
            UPDATE components
            SET inventory_id = (
                SELECT inventory.id
                FROM inventory
                WHERE inventory.name = components.name
                LIMIT 1
            )
            WHERE inventory_id IS NULL
        """)

        c.commit()

    log("Migration: component inventory_id verified.")

# --- EXPORT HELPERS (add to db_ops.py) ---

_ALLOWED_TABLES = {"items", "components", "recipes", "inventory"}

def _query_all(conn, table):
    return conn.execute(f"SELECT * FROM {table}").fetchall(), [d[0] for d in conn.execute(f"PRAGMA table_info({table})")]

def export_selected_as_zip_csv(tables: list[str]) -> bytes:
    """
    Return a ZIP (bytes) containing one CSV per selected table.
    """
    with connect() as c:
        mem = io.BytesIO()
        with zipfile.ZipFile(mem, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            for t in tables:
                if t not in _ALLOWED_TABLES: 
                    continue
                rows = c.execute(f"SELECT * FROM {t}").fetchall()
                cols = [col[1] for col in c.execute(f"PRAGMA table_info({t})")]
                # Write CSV to an in-memory buffer
                buf = io.StringIO()
                writer = csv.writer(buf)
                writer.writerow(cols)
                for r in rows:
                    writer.writerow([r[k] if isinstance(r, sqlite3.Row) else r[i] for i, k in enumerate(cols)])
                zf.writestr(f"{t}.csv", buf.getvalue())
        return mem.getvalue()

def export_selected_as_sql(tables: list[str]) -> str:
    """
    Return a SQL script that recreates ONLY the selected tables' schema + data.
    (This uses sqlite's schema + INSERTs for selected tables.)
    """
    with connect() as c:
        out = io.StringIO()
        # Write schema for selected tables
        for t in tables:
            if t not in _ALLOWED_TABLES:
                continue
            schema = c.execute(
                "SELECT sql FROM sqlite_master WHERE type='table' AND name=?", (t,)
            ).fetchone()
            if schema and schema["sql"]:
                out.write(schema["sql"] + ";\n\n")
        # Write INSERTs
        for t in tables:
            if t not in _ALLOWED_TABLES:
                continue
            rows = c.execute(f"SELECT * FROM {t}").fetchall()
            cols = [col[1] for col in c.execute(f"PRAGMA table_info({t})")]
            for r in rows:
                vals = []
                for k in cols:
                    v = r[k]
                    if v is None:
                        vals.append("NULL")
                    elif isinstance(v, (int, float)):
                        vals.append(str(v))
                    else:
                        # escape single quotes
                        s = str(v).replace("'", "''")
                        vals.append(f"'{s}'")
                out.write(f"INSERT INTO {t} ({', '.join(cols)}) VALUES ({', '.join(vals)});\n")
            out.write("\n")
        return out.getvalue()

CATALOG_TABLES = [
    "inventory",
    "items",
    "components",
    "recipes",
    "sales",
    "sale_requirements",
    "settings",
    "flavors",
]

FULL_BACKUP_TABLES = [
    "inventory",
    "items",
    "components",
    "recipes",
    "sales",
    "sale_requirements",
    "settings",
    "flavors",
    "admin_pins",
    "orders",
    "order_lines",
    "order_discounts",
    "shifts",
    "expenses",
]


def _table_exists(c, table_name: str) -> bool:
    row = c.execute("""
        SELECT name
        FROM sqlite_master
        WHERE type='table' AND name=?
    """, (table_name,)).fetchone()

    return row is not None


def _get_export_tables(export_type: str) -> list[str]:
    if export_type == "full":
        return FULL_BACKUP_TABLES

    return CATALOG_TABLES


def _exports_dir() -> str:
    base = os.path.join(os.getcwd(), "exports")
    os.makedirs(base, exist_ok=True)
    return base


def export_database(export_type: str = "catalog", fmt: str = "csv") -> str:
    """
    Export catalog setup or full database backup.

    export_type:
      catalog = setup data only
      full    = setup + operational records

    fmt:
      csv = ZIP of CSV files
      sql = one SQL dump file
    """
    export_type = export_type if export_type in ("catalog", "full") else "catalog"
    fmt = fmt if fmt in ("csv", "sql") else "csv"

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = _exports_dir()

    if fmt == "sql":
        filename = f"brewins_{export_type}_backup_{stamp}.sql"
        path = os.path.join(out_dir, filename)
        export_database_sql(path, export_type)
        return path

    filename = f"brewins_{export_type}_backup_{stamp}.zip"
    path = os.path.join(out_dir, filename)
    export_database_csv_zip(path, export_type)
    return path


def export_database_csv_zip(path: str, export_type: str = "catalog"):
    tables = _get_export_tables(export_type)

    with connect() as c:
        with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
            exported_tables = []

            for table in tables:
                if not _table_exists(c, table):
                    continue

                exported_tables.append(table)

                rows = c.execute(f'SELECT * FROM "{table}"').fetchall()
                cols = [r["name"] for r in c.execute(f'PRAGMA table_info("{table}")').fetchall()]

                csv_buffer = io.StringIO()
                writer = csv.writer(csv_buffer)

                writer.writerow(cols)

                for row in rows:
                    writer.writerow([row[col] for col in cols])

                zf.writestr(f"{table}.csv", csv_buffer.getvalue())

            manifest_buffer = io.StringIO()
            writer = csv.writer(manifest_buffer)

            writer.writerow(["field", "value"])
            writer.writerow(["export_type", export_type])
            writer.writerow(["created_at", datetime.now().isoformat(timespec="seconds")])
            writer.writerow(["format", "csv_zip"])
            writer.writerow(["tables", ", ".join(exported_tables)])

            zf.writestr("_manifest.csv", manifest_buffer.getvalue())


def export_database_sql(path: str, export_type: str = "catalog"):
    tables = _get_export_tables(export_type)

    with connect() as c:
        with open(path, "w", encoding="utf-8") as f:
            f.write("-- BrewIns POS Export\n")
            f.write(f"-- Export type: {export_type}\n")
            f.write(f"-- Created at: {datetime.now().isoformat(timespec='seconds')}\n\n")

            f.write("PRAGMA foreign_keys=OFF;\n")
            f.write("BEGIN TRANSACTION;\n\n")

            for table in tables:
                if not _table_exists(c, table):
                    continue

                schema = c.execute("""
                    SELECT sql
                    FROM sqlite_master
                    WHERE type='table' AND name=?
                """, (table,)).fetchone()

                if schema and schema["sql"]:
                    f.write(f'DROP TABLE IF EXISTS "{table}";\n')
                    f.write(schema["sql"] + ";\n\n")

                rows = c.execute(f'SELECT * FROM "{table}"').fetchall()
                cols = [r["name"] for r in c.execute(f'PRAGMA table_info("{table}")').fetchall()]

                for row in rows:
                    values = [row[col] for col in cols]
                    sql = _make_insert_sql(table, cols, values)
                    f.write(sql + "\n")

                f.write("\n")

            f.write("COMMIT;\n")
            f.write("PRAGMA foreign_keys=ON;\n")


def _make_insert_sql(table: str, cols: list[str], values: list) -> str:
    col_list = ", ".join([f'"{col}"' for col in cols])
    value_list = ", ".join([_sql_literal(v) for v in values])

    return f'INSERT INTO "{table}" ({col_list}) VALUES ({value_list});'


def _sql_literal(value):
    if value is None:
        return "NULL"

    if isinstance(value, (int, float)):
        return str(value)

    text = str(value).replace("'", "''")
    return f"'{text}'"

def _restore_uploads_dir() -> str:
    base = os.path.join(os.getcwd(), "restore_uploads")
    os.makedirs(base, exist_ok=True)
    return base


def _safety_backups_dir() -> str:
    base = os.path.join(os.getcwd(), "safety_backups")
    os.makedirs(base, exist_ok=True)
    return base

def restore_database_from_sql(sql_path: str, restore_type: str = "catalog") -> dict:
    """
    Restore database content from a SQL export.

    restore_type:
      catalog = restore setup/catalog tables only from SQL export
      full    = restore full SQL backup

    This assumes the SQL file came from our export system.
    """
    restore_type = restore_type if restore_type in ("catalog", "full") else "catalog"

    safety_backup = make_safety_db_backup()

    with open(sql_path, "r", encoding="utf-8") as f:
        sql_text = f.read()

    with connect() as c:
        if restore_type == "full":
            # Full restore runs the SQL dump as-is.
            c.executescript(sql_text)
            c.commit()

        else:
            # Catalog restore should only wipe/rebuild catalog tables.
            # Since our catalog SQL export only contains catalog tables,
            # running it is safe as long as the uploaded file is truly catalog export.
            c.executescript(sql_text)
            c.commit()

    return {
        "ok": True,
        "restore_type": restore_type,
        "restored_from": sql_path,
        "safety_backup": safety_backup
    }
# ---------------------------------------------------------------------
# Reports / KPIs
# ---------------------------------------------------------------------
def sales_summary():
    """Daily gross totals for /reports screen."""
    with connect() as c:
        rows = c.execute("""
            SELECT DATE(ts) AS date,
                   COUNT(*) AS orders,
                   ROUND(SUM(total), 2) AS gross
            FROM orders
            GROUP BY DATE(ts)
            ORDER BY date DESC
        """).fetchall()
        return [dict(r) for r in rows]

def kpis(start_date: str, end_date: str):
    """Detailed KPI block: gross sales, tax, COGS, and gross profit."""
    with connect() as c:
        orders = c.execute("""
            SELECT id, subtotal, discount, tax
            FROM orders
            WHERE DATE(ts) BETWEEN ? AND ?
        """, (start_date, end_date)).fetchall()

        gross = 0.0
        tax = 0.0
        cogs = 0.0

        for o in orders:
            gross += float(o["subtotal"] or 0) - float(o["discount"] or 0)
            tax += float(o["tax"] or 0)

            lines = c.execute("""
                SELECT item_id, qty
                FROM order_lines
                WHERE order_id = ?
            """, (o["id"],)).fetchall()

            for ln in lines:
                qty_sold = float(ln["qty"] or 0)

                recs = c.execute("""
                    SELECT
                        r.qty_per_item,
                        CASE
                            WHEN inv.units_per_case > 0
                            THEN inv.case_cost / inv.units_per_case
                            ELSE 0
                        END AS unit_cost
                    FROM recipes r
                    JOIN components comp ON comp.id = r.component_id
                    LEFT JOIN inventory inv ON inv.id = comp.inventory_id
                    WHERE r.item_id = ?
                """, (ln["item_id"],)).fetchall()

                for rr in recs:
                    qty_per_item = float(rr["qty_per_item"] or 0)
                    unit_cost = float(rr["unit_cost"] or 0)

                    cogs += qty_sold * qty_per_item * unit_cost

        gp = gross - cogs

        return {
            "gross": round(gross, 2),
            "tax": round(tax, 2),
            "cogs": round(cogs, 2),
            "gp": round(gp, 2)
        }
    
# --- REPORTS: Orders + Receipts ---

def list_orders_by_date(date_str: str):
    with connect() as c:
        rows = c.execute("""
            SELECT id, ts, cashier, method, subtotal, discount, tax, total,
                   COALESCE(receipt_seq, id) AS receipt_no
            FROM orders
            WHERE DATE(ts) = ?
            ORDER BY receipt_no ASC, id ASC
        """, (date_str,)).fetchall()
        return [dict(r) for r in rows]

def get_receipt(order_id: int):
    with connect() as c:
        order = c.execute("""
            SELECT id, ts, cashier, method,
                   subtotal,
                   discount,
                   discount_type,
                   tax,
                   total,
                   COALESCE(receipt_seq, id) AS receipt_no,
                   COALESCE(receipt_date, substr(ts,1,10)) AS receipt_date
            FROM orders
            WHERE id = ?
        """, (order_id,)).fetchone()

        if not order:
            return None
        lines = c.execute("""
            SELECT ol.id, ol.item_id, i.name AS item_name, ol.qty, ol.unit_price,
                   ROUND(ol.qty * ol.unit_price, 2) AS line_total
            FROM order_lines ol
            JOIN items i ON i.id = ol.item_id
            WHERE ol.order_id = ?
            ORDER BY ol.id ASC
        """, (order_id,)).fetchall()    
        line_dicts = []

        for ln in lines:
            line = dict(ln)
            modifier_rows = c.execute("""
                SELECT id,
                       modifier_type,
                       name,
                       qty,
                       unit_price,
                       line_total
                FROM order_line_modifiers
                WHERE order_line_id = ?
                ORDER BY id ASC
            """, (line["id"],)).fetchall()

            line["modifiers"] = [dict(m) for m in modifier_rows]
            line_dicts.append(line)

        discounts = c.execute("""
            SELECT id,
                sale_id,
                sale_name,
                sale_type,
                source,
                amount,
                authorized_pin_id,
                authorized_pin_label,
                created_at
            FROM order_discounts
            WHERE order_id = ?
            ORDER BY id ASC
        """, (order_id,)).fetchall()
        return {
            "order": dict(order),
            "lines": line_dicts,
            "discounts": [dict(r) for r in discounts]
        }
        


def _today_date(ts_iso: str | None = None) -> str:
    # ts_iso like '2025-10-11T14:22:00'; default now
    if not ts_iso:
        return datetime.now().date().isoformat()
    return ts_iso[:10]

def _next_receipt_seq(date_str: str) -> int:
    with connect() as c:
        row = c.execute("SELECT COALESCE(MAX(receipt_seq),0)+1 AS n FROM orders WHERE receipt_date = ?", (date_str,)).fetchone()
        return int(row["n"] or 1)


# ------- Delete Order and Restore Inventory -------
def delete_order_and_restore(order_id: int) -> bool:
    """
    Restores inventory quantities consumed by the order (per recipes),
    then deletes the order (order_lines are ON DELETE CASCADE).
    """
    with connect() as c:
        # 1) Gather usage by inventory_id for this order
        rows = c.execute("""
            SELECT inv.id AS inv_id,
                   inv.name AS inv_name,
                   SUM(ol.qty * r.qty_per_item) AS used_qty
            FROM order_lines ol
            JOIN recipes r       ON r.item_id = ol.item_id
            JOIN components comp ON comp.id   = r.component_id
            JOIN inventory inv   ON inv.id    = comp.inventory_id
            WHERE ol.order_id = ?
            GROUP BY inv.id, inv.name
        """, (order_id,)).fetchall()

        # 2) Restore inventory
        for row in rows:
            used = float(row["used_qty"] or 0)
            if used <= 0:
                continue

            c.execute("""
                UPDATE inventory
                SET qty_on_hand = qty_on_hand + ?
                WHERE id = ?
            """, (used, row["inv_id"]))

        # 3) Delete order
        c.execute("DELETE FROM orders WHERE id=?", (order_id,))
        c.commit()
        return True

# ------- Daily Sales Summary + Export to XLSX -------


def daily_sales_summary(date_str: str):
    """Return item sales, discounts, COGS, flavor add-ons, and profit summary for a given date."""
    with connect() as c:
        # Item-level sales before discounts.
        #
        # Important:
        #   i.price = base menu price
        #   ol.unit_price = configured sold price after modifiers/flavors
        #   order_line_modifiers = detail rows for flavor pump revenue
        #
        # We group by configured unit price so two different final prices for
        # the same base drink stay visible instead of getting averaged together.
        lines = c.execute("""
            SELECT
                i.id AS item_id,
                i.name AS item_name,
                i.price AS base_unit_price,
                ol.unit_price AS unit_price,
                SUM(ol.qty) AS qty_sold,
                ROUND(SUM(ol.qty * i.price), 2) AS base_gross,
                ROUND(SUM(ol.qty * ol.unit_price), 2) AS gross,
                ROUND(SUM(ol.qty * (ol.unit_price - i.price)), 2) AS modifier_gross
            FROM orders o
            JOIN order_lines ol ON o.id = ol.order_id
            JOIN items i ON i.id = ol.item_id
            WHERE DATE(o.ts) = ?
            GROUP BY i.id, i.name, i.price, ol.unit_price
            ORDER BY i.name, ol.unit_price
        """, (date_str,)).fetchall()

        items = []
        pre_discount_sales = 0.0
        base_sales_total = 0.0
        modifier_sales_total = 0.0
        cogs_total = 0.0

        for l in lines:
            item = dict(l)
            item_id = item["item_id"]
            qty_sold = float(item["qty_sold"] or 0)
            base_unit_price = float(item["base_unit_price"] or 0)
            unit_price = float(item["unit_price"] or 0)
            base_gross = float(item["base_gross"] or 0)
            modifier_gross = float(item["modifier_gross"] or 0)
            gross = float(item["gross"] or 0)

            comps = c.execute("""
                SELECT r.qty_per_item,
                       CASE WHEN inv.units_per_case > 0
                            THEN inv.case_cost / inv.units_per_case
                            ELSE 0 END AS unit_cost
                FROM recipes r
                JOIN components comp ON comp.id = r.component_id
                JOIN inventory inv ON inv.id = comp.inventory_id
                WHERE r.item_id = ?
            """, (item_id,)).fetchall()

            cogs = sum(
                qty_sold * float(r["qty_per_item"] or 0) * float(r["unit_cost"] or 0)
                for r in comps
            )

            items.append({
                "name": item["item_name"],
                "qty_sold": qty_sold,
                "base_unit_price": round(base_unit_price, 2),
                "unit_price": round(unit_price, 2),  # final configured price
                "modifier_unit_price": round(unit_price - base_unit_price, 2),
                "base_gross": round(base_gross, 2),
                "modifier_gross": round(modifier_gross, 2),
                "gross": round(gross, 2),
                "cogs": round(cogs, 2)
            })

            pre_discount_sales += gross
            base_sales_total += base_gross
            modifier_sales_total += modifier_gross
            cogs_total += cogs

        # Flavor/add-on breakdown by modifier name.
        modifier_rows = c.execute("""
            SELECT
                olm.modifier_type,
                olm.name,
                SUM(olm.qty * ol.qty) AS qty_sold,
                ROUND(SUM(olm.line_total * ol.qty), 2) AS gross
            FROM orders o
            JOIN order_lines ol ON o.id = ol.order_id
            JOIN order_line_modifiers olm ON olm.order_line_id = ol.id
            WHERE DATE(o.ts) = ?
            GROUP BY olm.modifier_type, olm.name
            ORDER BY olm.modifier_type, olm.name
        """, (date_str,)).fetchall()

        modifiers = [dict(r) for r in modifier_rows]

        # Order-level totals
        order_totals = c.execute("""
            SELECT
                COALESCE(SUM(subtotal), 0) AS subtotal_total,
                COALESCE(SUM(discount), 0) AS discount_total,
                COALESCE(SUM(tax), 0) AS tax_total,
                COALESCE(SUM(total), 0) AS net_sales
            FROM orders
            WHERE DATE(ts) = ?
        """, (date_str,)).fetchone()

        discount_total = float(order_totals["discount_total"] or 0)
        tax_total = float(order_totals["tax_total"] or 0)
        net_sales = float(order_totals["net_sales"] or 0)

        # Discount details by source/sale
        discount_rows = c.execute("""
            SELECT
                COALESCE(source, 'UNKNOWN') AS source,
                COALESCE(sale_name, sale_type, 'Discount') AS sale_name,
                COALESCE(SUM(amount), 0) AS amount
            FROM order_discounts od
            JOIN orders o ON o.id = od.order_id
            WHERE DATE(o.ts) = ?
            GROUP BY source, sale_name
            ORDER BY source, sale_name
        """, (date_str,)).fetchall()

        discounts = [dict(r) for r in discount_rows]

        # Fallback if old orders have orders.discount but no order_discounts rows
        detailed_discount_total = sum(float(d["amount"] or 0) for d in discounts)
        missing_discount = round(discount_total - detailed_discount_total, 2)

        if missing_discount > 0:
            discounts.append({
                "source": "LEGACY",
                "sale_name": "Older discount record",
                "amount": missing_discount
            })

        gp = net_sales - cogs_total

        exp = c.execute("""
            SELECT COALESCE(SUM(amount),0) AS total
            FROM expenses
            WHERE date=?
        """, (date_str,)).fetchone()["total"] or 0.0

        np = gp - float(exp)

        return {
            "date": date_str,
            "items": items,
            "modifiers": modifiers,

            "base_sales_total": round(base_sales_total, 2),
            "modifier_sales_total": round(modifier_sales_total, 2),
            "pre_discount_sales": round(pre_discount_sales, 2),
            "gross_total": round(pre_discount_sales, 2),  # keeps old frontend compatible

            "discount_total": round(discount_total, 2),
            "discounts": discounts,

            "tax_total": round(tax_total, 2),
            "net_sales": round(net_sales, 2),

            "cogs_total": round(cogs_total, 2),
            "gp": round(gp, 2),

            "expenses": round(float(exp), 2),
            "net_profit": round(np, 2)
        }


def export_sales_to_xlsx(date_str: str) -> bytes:
    """Return an Excel file (bytes) for that day’s sales summary."""
    data = daily_sales_summary(date_str)
    wb = Workbook()
    ws = wb.active
    ws.title = f"Sales {date_str}"

    ws.append(["Item", "Qty Sold", "Base Price", "Flavor/Add-on Per Item", "Final Unit Price", "Base Sales", "Flavor/Add-on Sales", "Pre-Discount Sales", "COGS"])
    for it in data["items"]:
        ws.append([
            it["name"],
            it["qty_sold"],
            it.get("base_unit_price", 0),
            it.get("modifier_unit_price", 0),
            it["unit_price"],
            it.get("base_gross", 0),
            it.get("modifier_gross", 0),
            it["gross"],
            it["cogs"]
        ])

    if data.get("modifiers"):
        ws.append([])
        ws.append(["Flavor/Add-on Breakdown"])
        ws.append(["Type", "Name", "Qty Sold", "Gross"])
        for m in data["modifiers"]:
            ws.append([
                m.get("modifier_type"),
                m.get("name"),
                m.get("qty_sold"),
                m.get("gross")
            ])

    ws.append([])
    ws.append(["Base Item Sales", data.get("base_sales_total", 0)])
    ws.append(["Flavor/Add-on Sales", data.get("modifier_sales_total", 0)])
    ws.append(["Pre-Discount Sales", data["pre_discount_sales"]])
    ws.append(["Discounts", -data["discount_total"]])

    if data.get("tax_total", 0) > 0:
        ws.append(["Tax", data["tax_total"]])

    ws.append(["Net Sales", data["net_sales"]])
    ws.append(["COGS Total", data["cogs_total"]])
    ws.append(["Gross Profit", data["gp"]])
    ws.append(["Operating Expenses", data["expenses"]])
    ws.append(["Net Profit", data["net_profit"]])

    if data.get("discounts"):
        ws.append([])
        ws.append(["Discount Details"])
        ws.append(["Source", "Sale", "Amount"])
        for d in data["discounts"]:
            ws.append([
                d.get("source"),
                d.get("sale_name"),
                -float(d.get("amount") or 0)
            ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

#------ Extended Sales Summary for Date Range + Export to XLSX -------


def sales_summary_range(start_date: str, end_date: str):
    """
    Aggregate item sales, discounts, COGS, flavor add-ons, and profit for DATE(ts)
    BETWEEN start_date AND end_date inclusive.
    """
    with connect() as c:
        rows = c.execute("""
            SELECT
                i.id AS item_id,
                i.name AS item_name,
                i.price AS base_unit_price,
                ol.unit_price AS unit_price,
                SUM(ol.qty) AS qty_sold,
                ROUND(SUM(ol.qty * i.price), 2) AS base_gross,
                ROUND(SUM(ol.qty * ol.unit_price), 2) AS gross,
                ROUND(SUM(ol.qty * (ol.unit_price - i.price)), 2) AS modifier_gross
            FROM orders o
            JOIN order_lines ol ON o.id = ol.order_id
            JOIN items i ON i.id = ol.item_id
            WHERE DATE(o.ts) BETWEEN ? AND ?
            GROUP BY i.id, i.name, i.price, ol.unit_price
            ORDER BY i.name, ol.unit_price
        """, (start_date, end_date)).fetchall()

        items = []
        pre_discount_sales = 0.0
        base_sales_total = 0.0
        modifier_sales_total = 0.0
        cogs_total = 0.0

        for r in rows:
            item_id = r["item_id"]
            qty_sold = float(r["qty_sold"] or 0)
            base_unit_price = float(r["base_unit_price"] or 0)
            unit_price = float(r["unit_price"] or 0)
            base_gross = float(r["base_gross"] or 0)
            modifier_gross = float(r["modifier_gross"] or 0)
            gross = float(r["gross"] or 0)

            comps = c.execute("""
                SELECT r.qty_per_item,
                       CASE WHEN inv.units_per_case > 0
                            THEN inv.case_cost / inv.units_per_case
                            ELSE 0 END AS unit_cost
                FROM recipes r
                JOIN components comp ON comp.id = r.component_id
                JOIN inventory inv ON inv.id = comp.inventory_id
                WHERE r.item_id = ?
            """, (item_id,)).fetchall()

            cogs = sum(
                qty_sold * float(x["qty_per_item"] or 0) * float(x["unit_cost"] or 0)
                for x in comps
            )

            items.append({
                "name": r["item_name"],
                "qty_sold": qty_sold,
                "base_unit_price": round(base_unit_price, 2),
                "unit_price": round(unit_price, 2),
                "modifier_unit_price": round(unit_price - base_unit_price, 2),
                "base_gross": round(base_gross, 2),
                "modifier_gross": round(modifier_gross, 2),
                "gross": round(gross, 2),
                "cogs": round(cogs, 2),
            })

            pre_discount_sales += gross
            base_sales_total += base_gross
            modifier_sales_total += modifier_gross
            cogs_total += cogs

        modifier_rows = c.execute("""
            SELECT
                olm.modifier_type,
                olm.name,
                SUM(olm.qty * ol.qty) AS qty_sold,
                ROUND(SUM(olm.line_total * ol.qty), 2) AS gross
            FROM orders o
            JOIN order_lines ol ON o.id = ol.order_id
            JOIN order_line_modifiers olm ON olm.order_line_id = ol.id
            WHERE DATE(o.ts) BETWEEN ? AND ?
            GROUP BY olm.modifier_type, olm.name
            ORDER BY olm.modifier_type, olm.name
        """, (start_date, end_date)).fetchall()

        modifiers = [dict(r) for r in modifier_rows]

        order_totals = c.execute("""
            SELECT
                COALESCE(SUM(subtotal), 0) AS subtotal_total,
                COALESCE(SUM(discount), 0) AS discount_total,
                COALESCE(SUM(tax), 0) AS tax_total,
                COALESCE(SUM(total), 0) AS net_sales
            FROM orders
            WHERE DATE(ts) BETWEEN ? AND ?
        """, (start_date, end_date)).fetchone()

        discount_total = float(order_totals["discount_total"] or 0)
        tax_total = float(order_totals["tax_total"] or 0)
        net_sales = float(order_totals["net_sales"] or 0)

        discount_rows = c.execute("""
            SELECT
                COALESCE(source, 'UNKNOWN') AS source,
                COALESCE(sale_name, sale_type, 'Discount') AS sale_name,
                COALESCE(SUM(amount), 0) AS amount
            FROM order_discounts od
            JOIN orders o ON o.id = od.order_id
            WHERE DATE(o.ts) BETWEEN ? AND ?
            GROUP BY source, sale_name
            ORDER BY source, sale_name
        """, (start_date, end_date)).fetchall()

        discounts = [dict(r) for r in discount_rows]

        detailed_discount_total = sum(float(d["amount"] or 0) for d in discounts)
        missing_discount = round(discount_total - detailed_discount_total, 2)

        if missing_discount > 0:
            discounts.append({
                "source": "LEGACY",
                "sale_name": "Older discount record",
                "amount": missing_discount
            })

        gp = net_sales - cogs_total

        exp = c.execute("""
            SELECT COALESCE(SUM(amount),0) AS total
            FROM expenses
            WHERE date BETWEEN ? AND ?
        """, (start_date, end_date)).fetchone()["total"] or 0.0

        np = gp - float(exp)

        return {
            "start": start_date,
            "end": end_date,
            "items": items,
            "modifiers": modifiers,

            "base_sales_total": round(base_sales_total, 2),
            "modifier_sales_total": round(modifier_sales_total, 2),
            "pre_discount_sales": round(pre_discount_sales, 2),
            "gross_total": round(pre_discount_sales, 2),

            "discount_total": round(discount_total, 2),
            "discounts": discounts,

            "tax_total": round(tax_total, 2),
            "net_sales": round(net_sales, 2),

            "cogs_total": round(cogs_total, 2),
            "gp": round(gp, 2),

            "expenses": round(float(exp), 2),
            "net_profit": round(np, 2),
        }

def export_sales_range_to_xlsx(start_date: str, end_date: str) -> bytes:
    data = sales_summary_range(start_date, end_date)
    wb = Workbook()

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append([f"Sales Summary {data['start']} to {data['end']}"])
    ws1.append([])
    ws1.append(["Base Item Sales", data.get("base_sales_total", 0)])
    ws1.append(["Flavor/Add-on Sales", data.get("modifier_sales_total", 0)])
    ws1.append(["Pre-Discount Sales", data["pre_discount_sales"]])
    ws1.append(["Discounts", -data["discount_total"]])

    if data.get("tax_total", 0) > 0:
        ws1.append(["Tax", data["tax_total"]])

    ws1.append(["Net Sales", data["net_sales"]])
    ws1.append(["COGS Total", data["cogs_total"]])
    ws1.append(["Gross Profit", data["gp"]])
    ws1.append(["Operating Expenses", data["expenses"]])
    ws1.append(["Net Profit", data["net_profit"]])

    if data.get("discounts"):
        ws1.append([])
        ws1.append(["Discount Details"])
        ws1.append(["Source", "Sale", "Amount"])
        for d in data["discounts"]:
            ws1.append([
                d.get("source"),
                d.get("sale_name"),
                -float(d.get("amount") or 0)
            ])

    # Sheet 2: By Item
    ws2 = wb.create_sheet("By Item")
    ws2.append(["Item", "Qty Sold", "Base Price", "Flavor/Add-on Per Item", "Final Unit Price", "Base Sales", "Flavor/Add-on Sales", "Pre-Discount Sales", "COGS"])
    for it in data["items"]:
        ws2.append([
            it["name"],
            it["qty_sold"],
            it.get("base_unit_price", 0),
            it.get("modifier_unit_price", 0),
            it["unit_price"],
            it.get("base_gross", 0),
            it.get("modifier_gross", 0),
            it["gross"],
            it["cogs"]
        ])

    if data.get("modifiers"):
        ws3 = wb.create_sheet("Flavor Add-ons")
        ws3.append(["Type", "Name", "Qty Sold", "Gross"])
        for m in data["modifiers"]:
            ws3.append([
                m.get("modifier_type"),
                m.get("name"),
                m.get("qty_sold"),
                m.get("gross")
            ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ---------------------------------------------------------------------
# Shift summaries for admin reports
# ---------------------------------------------------------------------
def list_shifts(start_date: str = None, end_date: str = None) -> list[dict]:
    """Return shift history with key financial details."""
    with connect() as c:
        if start_date and end_date:
            rows = c.execute("""
                SELECT id,
                       DATE(ts_start) AS date,
                       cashier,
                       opening_float,
                       closing_amount,
                       over_short,
                       ROUND((closing_amount - opening_float), 2) AS net_change,
                       ts_start,
                       ts_end,
                       is_active
                FROM shifts
                WHERE DATE(ts_start) BETWEEN ? AND ?
                ORDER BY ts_start DESC
            """, (start_date, end_date)).fetchall()
        else:
            rows = c.execute("""
                SELECT id,
                       DATE(ts_start) AS date,
                       cashier,
                       opening_float,
                       closing_amount,
                       over_short,
                       ROUND((closing_amount - opening_float), 2) AS net_change,
                       ts_start,
                       ts_end,
                       is_active
                FROM shifts
                ORDER BY ts_start DESC
            """).fetchall()
        return [dict(r) for r in rows]

# ---------------------------------------------------------------------
# Settings and Discount Pins
# ---------------------------------------------------------------------
def get_setting(key: str, default=None):
    with connect() as c:
        row = c.execute(
            "SELECT value FROM settings WHERE key=?",
            (key,)
        ).fetchone()

        return row["value"] if row else default


def set_setting(key: str, value):
    with connect() as c:
        c.execute("""
            INSERT INTO settings(key, value)
            VALUES (?, ?)
            ON CONFLICT(key)
            DO UPDATE SET value=excluded.value
        """, (key, str(value)))
        c.commit()


def get_pos_settings():
    return {
        "require_pin_discounts": get_setting("require_pin_discounts", "0") == "1",
        "require_pin_admin_access": get_setting("require_pin_admin_access", "0") == "1",
        "require_pin_exit_program": get_setting("require_pin_exit_program", "0") == "1",
        "tax_rate": float(get_setting("tax_rate", "0") or 0),
    }


def list_admin_pins():
    with connect() as c:
        rows = c.execute("""
            SELECT id, label, active
            FROM admin_pins
            ORDER BY active DESC, id ASC
        """).fetchall()

        return [dict(r) for r in rows]


def add_admin_pin(pin: str, label: str = ""):
    with connect() as c:
        c.execute("""
            INSERT INTO admin_pins(pin, label, active)
            VALUES (?, ?, 1)
        """, (pin.strip(), label.strip()))
        c.commit()


def delete_admin_pin(pin_id: int):
    with connect() as c:
        c.execute(
            "DELETE FROM admin_pins WHERE id=?",
            (pin_id,)
        )
        c.commit()


def validate_admin_pin(pin: str) -> bool:
    with connect() as c:
        row = c.execute("""
            SELECT id
            FROM admin_pins
            WHERE pin=? AND active=1
            LIMIT 1
        """, (pin.strip(),)).fetchone()

        return row is not None

def validate_admin_pin_info(pin: str):
    """
    Return PIN info if valid, otherwise None.
    Used when we need to record which labeled PIN authorized a manual discount.
    """
    with connect() as c:
        row = c.execute("""
            SELECT id, label
            FROM admin_pins
            WHERE pin=? AND active=1
            LIMIT 1
        """, (pin.strip(),)).fetchone()

        return dict(row) if row else None
    
def has_admin_pins() -> bool:
    """Return True if at least one active admin PIN exists."""
    with connect() as c:
        row = c.execute("""
            SELECT id
            FROM admin_pins
            WHERE active = 1
            LIMIT 1
        """).fetchone()

        return row is not None
